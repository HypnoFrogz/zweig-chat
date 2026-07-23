"""Nebenan Messenger — Admin panel (user management + feedback)."""

import csv
import io
import re
import uuid

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from helpers import get_current_user, get_display_name, require_admin, hash_password, now_iso
from database import get_db

router = APIRouter(prefix="/api/admin", tags=["admin"])

# Login is used both for auth and for @mentions in chat. Chat highlights
# mentions via /@(\w+)/, which only matches latin letters, digits and "_",
# so logins are constrained to that set. Display names stay free-form (Cyrillic ok).
USERNAME_RE = re.compile(r"^[a-z0-9_]{2,32}$")


def _normalize_username(raw) -> str:
    return str(raw or "").strip().lower()


def _validate_username(username: str) -> str | None:
    """Return an error message if invalid, else None."""
    if not username:
        return "Логин обязателен"
    if not USERNAME_RE.fullmatch(username):
        return "Логин: 2–32 символа, только латинские буквы, цифры и _"
    return None


async def _create_user_row(db, username: str, password: str, display_name: str, role: str):
    """Create a single user. Returns (ok: bool, detail: str).

    Does not commit — the caller commits once for the whole batch.
    """
    username = _normalize_username(username)
    password = str(password or "")
    display_name = str(display_name or "").strip()
    if role not in ("user", "admin"):
        role = "user"

    err = _validate_username(username)
    if err:
        return False, err
    if not password or len(password) < 3:
        return False, "Пароль слишком короткий (мин. 3 символа)"
    if not display_name:
        display_name = username.capitalize()

    c = await db.execute("SELECT 1 FROM users WHERE username = ?", (username,))
    if await c.fetchone():
        return False, "Логин уже существует"

    await db.execute(
        "INSERT INTO users (username, password, display_name, role, created_at) VALUES (?, ?, ?, ?, ?)",
        (username, hash_password(password), display_name, role, now_iso()),
    )
    return True, "created"

async def _revoke_user_sessions(db, target: str, reason: str):
    await db.execute(
        "UPDATE user_sessions SET revoked_at = ?, revoked_reason = ? WHERE username = ? AND revoked_at IS NULL",
        (now_iso(), reason, target),
    )


@router.get("/users")
async def list_all_users(username: str = Depends(get_current_user)):
    await require_admin(username)
    db = await get_db()
    cursor = await db.execute(
        "SELECT username, display_name, nickname, avatar_path, status_text, role, blocked, created_at, last_seen FROM users ORDER BY created_at"
    )
    return [dict(r) for r in await cursor.fetchall()]


@router.post("/users")
async def create_user(data: dict, username: str = Depends(get_current_user)):
    await require_admin(username)

    new_username = _normalize_username(data.get("username"))
    display_name = str(data.get("display_name") or "").strip() or new_username.capitalize()
    role = data.get("role", "user")

    db = await get_db()
    ok, detail = await _create_user_row(db, new_username, data.get("password", ""), display_name, role)
    if not ok:
        # 409 for duplicates, 400 for validation problems.
        status = 409 if "существует" in detail else 400
        raise HTTPException(status_code=status, detail=detail)

    await db.commit()
    return {
        "username": new_username,
        "display_name": display_name if display_name else new_username.capitalize(),
        "role": role if role in ("user", "admin") else "user",
        "created_at": now_iso(),
    }


# ── Bulk import from Excel / CSV ──────────────────────────────────

# Header synonyms → canonical field. Matched case-insensitively after strip.
_COLUMN_ALIASES = {
    "username": {"username", "login", "user", "логин", "юзернейм", "ник"},
    "password": {"password", "pass", "пароль"},
    "display_name": {"display_name", "display name", "name", "fio", "имя",
                     "отображаемое имя", "фио", "имя пользователя"},
    "role": {"role", "роль"},
}


def _map_headers(header_row) -> dict:
    """Map column index → canonical field name based on the header row."""
    mapping = {}
    for idx, cell in enumerate(header_row):
        key = str(cell or "").strip().lower()
        for field, aliases in _COLUMN_ALIASES.items():
            if key in aliases:
                mapping[idx] = field
                break
    return mapping


def _rows_from_upload(filename: str, content: bytes):
    """Yield lists of cell values for each row (including the header)."""
    name = (filename or "").lower()
    if name.endswith(".csv"):
        text = content.decode("utf-8-sig", errors="replace")
        # Sniff delimiter (comma or semicolon — Excel-RU often uses ';').
        sample = text[:2048]
        delimiter = ";" if sample.count(";") > sample.count(",") else ","
        for row in csv.reader(io.StringIO(text), delimiter=delimiter):
            yield row
        return
    # Default: .xlsx via openpyxl
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl не установлен на сервере")
    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="Не удалось прочитать файл. Поддерживаются .xlsx и .csv")
    ws = wb.active
    for row in ws.iter_rows(values_only=True):
        yield list(row)


@router.post("/users/import")
async def import_users(file: UploadFile = File(...), username: str = Depends(get_current_user)):
    await require_admin(username)

    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="Файл пуст")

    rows = list(_rows_from_upload(file.filename, content))
    if not rows:
        raise HTTPException(status_code=400, detail="В файле нет данных")

    mapping = _map_headers(rows[0])
    if "username" not in mapping.values() or "password" not in mapping.values():
        raise HTTPException(
            status_code=400,
            detail="Не найдены обязательные столбцы «Логин» и «Пароль». Скачайте шаблон.",
        )

    def cell(row, field):
        for idx, f in mapping.items():
            if f == field and idx < len(row):
                return row[idx]
        return ""

    db = await get_db()
    created, skipped, errors = [], [], []
    for line_no, row in enumerate(rows[1:], start=2):
        if not any(str(c or "").strip() for c in row):
            continue  # skip blank lines
        uname = _normalize_username(cell(row, "username"))
        ok, detail = await _create_user_row(
            db, uname, cell(row, "password"), cell(row, "display_name"), cell(row, "role") or "user"
        )
        if ok:
            created.append(uname)
        elif "существует" in detail:
            skipped.append(uname)
        else:
            errors.append({"line": line_no, "username": uname, "error": detail})

    await db.commit()
    return {
        "created": created,
        "skipped": skipped,
        "errors": errors,
        "created_count": len(created),
        "skipped_count": len(skipped),
        "error_count": len(errors),
    }


@router.get("/users/import/template")
async def import_template(username: str = Depends(get_current_user)):
    await require_admin(username)

    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Пользователи"
    ws.append(["Логин", "Пароль", "Отображаемое имя", "Роль"])
    ws.append(["ivan", "changeme123", "Иван Ковач", "user"])
    ws.append(["anna_k", "changeme123", "Анна", "user"])
    for col, width in zip("ABCD", (18, 18, 26, 10)):
        ws.column_dimensions[col].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="nebenan_users_template.xlsx"'},
    )


@router.put("/users/{target}/block")
async def toggle_block(target: str, data: dict, username: str = Depends(get_current_user)):
    await require_admin(username)

    if target == username:
        raise HTTPException(status_code=400, detail="Cannot block yourself")

    blocked = int(bool(data.get("blocked", True)))
    db = await get_db()

    c = await db.execute("SELECT 1 FROM users WHERE username = ?", (target,))
    if not await c.fetchone():
        raise HTTPException(status_code=404, detail="User not found")

    await db.execute("UPDATE users SET blocked = ? WHERE username = ?", (blocked, target))
    if blocked:
        await _revoke_user_sessions(db, target, "blocked_by_admin")
    await db.commit()
    return {"username": target, "blocked": bool(blocked)}


@router.put("/users/{target}/reset-password")
async def reset_password(target: str, data: dict, username: str = Depends(get_current_user)):
    await require_admin(username)

    new_password = data.get("password", "")
    if not new_password or len(new_password) < 3:
        raise HTTPException(status_code=400, detail="Password too short (min 3 chars)")

    db = await get_db()
    c = await db.execute("SELECT 1 FROM users WHERE username = ?", (target,))
    if not await c.fetchone():
        raise HTTPException(status_code=404, detail="User not found")

    hashed = hash_password(new_password)
    await db.execute("UPDATE users SET password = ? WHERE username = ?", (hashed, target))
    await _revoke_user_sessions(db, target, "password_reset_by_admin")
    await db.commit()
    return {"ok": True}


@router.put("/users/{target}/role")
async def change_role(target: str, data: dict, username: str = Depends(get_current_user)):
    await require_admin(username)

    if target == username:
        raise HTTPException(status_code=400, detail="Cannot change your own role")

    new_role = data.get("role", "user")
    if new_role not in ("user", "admin"):
        raise HTTPException(status_code=400, detail="Invalid role")

    db = await get_db()
    c = await db.execute("SELECT 1 FROM users WHERE username = ?", (target,))
    if not await c.fetchone():
        raise HTTPException(status_code=404, detail="User not found")

    await db.execute("UPDATE users SET role = ? WHERE username = ?", (new_role, target))
    await db.commit()
    return {"username": target, "role": new_role}


@router.delete("/users/{target}")
async def delete_user(target: str, username: str = Depends(get_current_user)):
    await require_admin(username)

    if target == username:
        raise HTTPException(status_code=400, detail="Cannot delete yourself")

    db = await get_db()
    c = await db.execute("SELECT 1 FROM users WHERE username = ?", (target,))
    if not await c.fetchone():
        raise HTTPException(status_code=404, detail="User not found")

    await db.execute("DELETE FROM users WHERE username = ?", (target,))
    await _revoke_user_sessions(db, target, "user_deleted")
    await db.commit()
    return {"ok": True}


# ── Feedback ──────────────────────────────────────────────────────

@router.post("/feedback")
async def submit_feedback(data: dict, username: str = Depends(get_current_user)):
    """Submit feedback (any authenticated user)."""
    text = (data.get("text") or "").strip()
    if not text:
        raise HTTPException(status_code=400, detail="Text is required")

    db = await get_db()
    display_name = await get_display_name(username)
    feedback_id = str(uuid.uuid4())
    now = now_iso()

    await db.execute(
        "INSERT INTO feedback (id, username, display_name, text, created_at) VALUES (?, ?, ?, ?, ?)",
        (feedback_id, username, display_name, text, now),
    )
    await db.commit()
    return {"ok": True}


@router.get("/feedback")
async def list_feedback(username: str = Depends(get_current_user)):
    """List all feedback (admin only)."""
    await require_admin(username)
    db = await get_db()
    cursor = await db.execute(
        "SELECT id, username, display_name, text, created_at FROM feedback ORDER BY created_at DESC"
    )
    return [dict(r) for r in await cursor.fetchall()]
